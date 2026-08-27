# -*- coding: utf-8 -*-
# 📂 apps/whatsapp_service/contacts_bulk.py
"""
سوق محجوب أونلاين - مسارات إدارة جهات الاتصال والرسائل الجماعية
Contacts Bulk Management Routes
"""

from flask import Blueprint, request, jsonify, render_template, current_app
import logging
import traceback
import os
import csv
import io
from datetime import datetime
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ✅ تعريف Blueprint خاص بجهات الاتصال
contacts_bulk_bp = Blueprint(
    'contacts_bulk',
    __name__,
    template_folder='templates',
    url_prefix='/admin/whatsapp'
)


# =========================================================================
# 1. عرض صفحة جهات الاتصال والإرسال الجماعي
# =========================================================================

@contacts_bulk_bp.route('/contacts-bulk', methods=['GET'])
def contacts_bulk_view():
    """عرض صفحة إدارة جهات الاتصال والرسائل الجماعية"""
    try:
        from apps.whatsapp_service.service import WhatsAppService
        wa_service = WhatsAppService()
        
        # جلب جميع جهات الاتصال
        contacts = wa_service.get_all_contacts()
        
        # إحصائيات الفئات
        stats = {
            'customers_count': len([c for c in contacts if c.get('category') == 'customers']),
            'merchants_count': len([c for c in contacts if c.get('category') == 'merchants']),
            'suppliers_count': len([c for c in contacts if c.get('category') == 'suppliers']),
            'marketers_count': len([c for c in contacts if c.get('category') == 'marketers']),
        }
        
        return render_template(
            'admin/contacts_bulk.html',
            contacts=contacts,
            stats=stats,
            current_category=request.args.get('category', 'all')
        )
    except Exception as e:
        logger.error(f"خطأ في عرض جهات الاتصال: {str(e)}")
        logger.error(traceback.format_exc())
        return render_template(
            'admin/contacts_bulk.html',
            contacts=[],
            stats={},
            error=str(e)
        ), 200


# =========================================================================
# 2. مسارات API لجهات الاتصال
# =========================================================================

@contacts_bulk_bp.route('/api/contacts', methods=['GET'])
def get_contacts_api():
    """جلب جميع جهات الاتصال (API)"""
    try:
        from apps.whatsapp_service.service import WhatsAppService
        wa_service = WhatsAppService()
        
        contacts = wa_service.get_all_contacts()
        return jsonify({"success": True, "contacts": contacts, "count": len(contacts)}), 200
    except Exception as e:
        logger.error(f"خطأ في جلب جهات الاتصال: {str(e)}")
        return jsonify({"success": False, "error": str(e), "contacts": []}), 500


@contacts_bulk_bp.route('/api/contacts/<int:contact_id>', methods=['GET'])
def get_contact_api(contact_id):
    """جلب جهة اتصال محددة"""
    try:
        from apps.models.whatsapp_models import WhatsAppCustomerContact
        from apps.extensions import db
        
        contact = WhatsAppCustomerContact.query.get(contact_id)
        if not contact:
            return jsonify({"success": False, "error": "جهة الاتصال غير موجودة"}), 404
        
        return jsonify({
            "success": True,
            "contact": {
                'id': contact.id,
                'name': contact.name,
                'phone': contact.phone,
                'category': getattr(contact, 'category', 'customers'),
                'city': getattr(contact, 'city', ''),
                'company': getattr(contact, 'company', ''),
                'email': getattr(contact, 'email', ''),
                'notes': getattr(contact, 'notes', ''),
                'unread_count': contact.unread_count or 0,
                'last_interaction': contact.last_timestamp.strftime('%Y-%m-%d %H:%M') if contact.last_timestamp else None
            }
        }), 200
    except Exception as e:
        logger.error(f"خطأ في جلب جهة الاتصال: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@contacts_bulk_bp.route('/api/contacts/add', methods=['POST'])
def add_contact_api():
    """إضافة جهة اتصال جديدة"""
    try:
        from apps.whatsapp_service.service import WhatsAppService
        wa_service = WhatsAppService()
        
        data = request.get_json(silent=True) or {}
        
        name = data.get('name', '').strip()
        phone = data.get('phone', '').strip()
        category = data.get('category', 'customers')
        city = data.get('city', '')
        company = data.get('company', '')
        email = data.get('email', '')
        notes = data.get('notes', '')
        
        if not name or not phone:
            return jsonify({"success": False, "error": "الاسم ورقم الهاتف مطلوبان"}), 400
        
        result = wa_service.add_contact(name, phone, category, city, company, email, notes)
        return jsonify(result), 200 if result.get('success') else 400
    except Exception as e:
        logger.error(f"خطأ في إضافة جهة الاتصال: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@contacts_bulk_bp.route('/api/contacts/<int:contact_id>', methods=['PUT'])
def update_contact_api(contact_id):
    """تحديث جهة اتصال"""
    try:
        from apps.whatsapp_service.service import WhatsAppService
        wa_service = WhatsAppService()
        
        data = request.get_json(silent=True) or {}
        result = wa_service.update_contact(contact_id, data)
        return jsonify(result), 200 if result.get('success') else 400
    except Exception as e:
        logger.error(f"خطأ في تحديث جهة الاتصال: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@contacts_bulk_bp.route('/api/contacts/<int:contact_id>', methods=['DELETE'])
def delete_contact_api(contact_id):
    """حذف جهة اتصال"""
    try:
        from apps.whatsapp_service.service import WhatsAppService
        wa_service = WhatsAppService()
        
        result = wa_service.delete_contact(contact_id)
        return jsonify(result), 200 if result.get('success') else 400
    except Exception as e:
        logger.error(f"خطأ في حذف جهة الاتصال: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@contacts_bulk_bp.route('/api/contacts/bulk-delete', methods=['POST'])
def bulk_delete_contacts_api():
    """حذف مجموعة من جهات الاتصال دفعة واحدة"""
    try:
        from apps.whatsapp_service.service import WhatsAppService
        wa_service = WhatsAppService()
        
        data = request.get_json(silent=True) or {}
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({"success": False, "error": "لم يتم تحديد أي جهات اتصال"}), 400
        
        result = wa_service.delete_contacts_bulk(ids)
        return jsonify(result), 200 if result.get('success') else 400
    except Exception as e:
        logger.error(f"خطأ في الحذف الجماعي: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


# =========================================================================
# 3. استيراد جهات الاتصال (بدون pandas)
# =========================================================================

@contacts_bulk_bp.route('/api/contacts/import', methods=['POST'])
def import_contacts_api():
    """استيراد جهات الاتصال من ملف CSV/Excel"""
    try:
        from apps.models.whatsapp_models import WhatsAppCustomerContact
        from apps.extensions import db
        
        if 'file' not in request.files:
            return jsonify({"success": False, "error": "لم يتم رفع ملف"}), 400
        
        file = request.files['file']
        default_category = request.form.get('default_category', 'customers')
        
        imported = 0
        errors = []
        
        # ✅ معالجة ملف CSV
        if file.filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
            csv_input = csv.DictReader(stream)
            for row in csv_input:
                try:
                    name = row.get('Name', row.get('name', '')).strip()
                    phone = row.get('Phone', row.get('phone', '')).strip()
                    
                    if not name or not phone:
                        continue
                    
                    # تنظيف رقم الهاتف
                    phone = phone.replace('+', '').replace(' ', '').strip()
                    
                    existing = WhatsAppCustomerContact.query.filter_by(phone=phone).first()
                    if existing:
                        continue
                    
                    contact = WhatsAppCustomerContact(
                        name=name,
                        phone=phone,
                        category=row.get('Category', row.get('category', default_category)),
                        city=row.get('City', row.get('city', '')),
                        company=row.get('Company', row.get('company', '')),
                        email=row.get('Email', row.get('email', '')),
                        notes=row.get('Notes', row.get('notes', '')),
                        unread_count=0,
                        last_timestamp=datetime.utcnow()
                    )
                    db.session.add(contact)
                    imported += 1
                except Exception as e:
                    errors.append(str(e))
                    continue
        
        # ✅ معالجة ملف Excel
        elif file.filename.endswith(('.xlsx', '.xls')):
            workbook = load_workbook(file)
            sheet = workbook.active
            
            # قراءة العناوين من الصف الأول
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # قراءة البيانات من الصف الثاني فما فوق
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    row_dict = dict(zip(headers, row))
                    
                    name = str(row_dict.get('Name', row_dict.get('name', ''))).strip()
                    phone = str(row_dict.get('Phone', row_dict.get('phone', ''))).strip()
                    
                    if not name or not phone:
                        continue
                    
                    # تنظيف رقم الهاتف
                    phone = phone.replace('+', '').replace(' ', '').strip()
                    
                    existing = WhatsAppCustomerContact.query.filter_by(phone=phone).first()
                    if existing:
                        continue
                    
                    contact = WhatsAppCustomerContact(
                        name=name,
                        phone=phone,
                        category=row_dict.get('Category', row_dict.get('category', default_category)),
                        city=str(row_dict.get('City', row_dict.get('city', ''))),
                        company=str(row_dict.get('Company', row_dict.get('company', ''))),
                        email=str(row_dict.get('Email', row_dict.get('email', ''))),
                        notes=str(row_dict.get('Notes', row_dict.get('notes', ''))),
                        unread_count=0,
                        last_timestamp=datetime.utcnow()
                    )
                    db.session.add(contact)
                    imported += 1
                except Exception as e:
                    errors.append(str(e))
                    continue
        
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "imported": imported, 
            "errors": errors,
            "message": f"تم استيراد {imported} جهة اتصال" + (f" مع {len(errors)} خطأ" if errors else "")
        }), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"خطأ في استيراد جهات الاتصال: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


# =========================================================================
# 4. تصدير جهات الاتصال المحددة
# =========================================================================

@contacts_bulk_bp.route('/api/contacts/export-selected', methods=['POST'])
def export_selected_contacts_api():
    """تصدير جهات الاتصال المحددة"""
    try:
        from apps.models.whatsapp_models import WhatsAppCustomerContact
        from apps.extensions import db
        
        data = request.get_json(silent=True) or {}
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({"success": False, "error": "لم يتم تحديد أي جهات اتصال"}), 400
        
        contacts = WhatsAppCustomerContact.query.filter(WhatsAppCustomerContact.id.in_(ids)).all()
        
        result = []
        for c in contacts:
            result.append({
                "id": c.id,
                "name": c.name,
                "phone": c.phone,
                "category": getattr(c, 'category', 'customers'),
                "city": getattr(c, 'city', ''),
                "company": getattr(c, 'company', ''),
                "email": getattr(c, 'email', ''),
                "notes": getattr(c, 'notes', ''),
                "unread_count": c.unread_count or 0,
                "last_interaction": c.last_timestamp.strftime('%Y-%m-%d %H:%M') if c.last_timestamp else None
            })
        
        return jsonify({"success": True, "data": result, "count": len(result)}), 200
    except Exception as e:
        logger.error(f"خطأ في تصدير جهات الاتصال: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@contacts_bulk_bp.route('/api/contacts/export-all', methods=['GET'])
def export_all_contacts_api():
    """تصدير جميع جهات الاتصال كملف JSON"""
    try:
        from apps.models.whatsapp_models import WhatsAppCustomerContact
        from apps.extensions import db
        
        contacts = WhatsAppCustomerContact.query.all()
        
        result = []
        for c in contacts:
            result.append({
                "id": c.id,
                "name": c.name,
                "phone": c.phone,
                "category": getattr(c, 'category', 'customers'),
                "city": getattr(c, 'city', ''),
                "company": getattr(c, 'company', ''),
                "email": getattr(c, 'email', ''),
                "notes": getattr(c, 'notes', ''),
                "unread_count": c.unread_count or 0,
                "last_interaction": c.last_timestamp.strftime('%Y-%m-%d %H:%M') if c.last_timestamp else None
            })
        
        return jsonify({"success": True, "data": result, "count": len(result)}), 200
    except Exception as e:
        logger.error(f"خطأ في تصدير جميع جهات الاتصال: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


# =========================================================================
# 5. مسارات الرسائل الجماعية والحملات
# =========================================================================

@contacts_bulk_bp.route('/api/contacts/send-bulk', methods=['POST'])
def send_bulk_messages_api():
    """إرسال رسائل جماعية لمجموعة من جهات الاتصال"""
    try:
        from apps.models.whatsapp_models import WhatsAppCustomerContact
        from apps.whatsapp_service.service import WhatsAppService
        from apps.extensions import db
        
        data = request.get_json(silent=True) or {}
        ids = data.get('ids', [])
        message = data.get('message', '').strip()
        
        if not ids:
            return jsonify({"success": False, "error": "لم يتم تحديد أي جهات اتصال"}), 400
        if not message:
            return jsonify({"success": False, "error": "الرسالة مطلوبة"}), 400
        
        contacts = WhatsAppCustomerContact.query.filter(WhatsAppCustomerContact.id.in_(ids)).all()
        wa_service = WhatsAppService()
        
        sent = 0
        failed = 0
        failed_list = []
        
        for contact in contacts:
            try:
                # تخصيص الرسالة لكل جهة
                msg = message.replace('{name}', contact.name or '')
                msg = msg.replace('{phone}', contact.phone or '')
                msg = msg.replace('{company}', getattr(contact, 'company', '') or '')
                msg = msg.replace('{city}', getattr(contact, 'city', '') or '')
                msg = msg.replace('{email}', getattr(contact, 'email', '') or '')
                
                wa_service.send_message(contact.phone, msg)
                sent += 1
            except Exception as e:
                logger.error(f"فشل إرسال رسالة إلى {contact.phone}: {str(e)}")
                failed += 1
                failed_list.append(contact.phone)
        
        return jsonify({
            "success": True,
            "sent": sent,
            "failed": failed,
            "failed_list": failed_list,
            "total": len(contacts),
            "message": f"تم إرسال {sent} رسالة بنجاح، فشل {failed}"
        }), 200
    except Exception as e:
        logger.error(f"خطأ في إرسال الرسائل الجماعية: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@contacts_bulk_bp.route('/api/campaigns/send', methods=['POST'])
def send_campaign_api():
    """إرسال حملة جماعية لفئة معينة"""
    try:
        from apps.models.whatsapp_models import WhatsAppCustomerContact
        from apps.whatsapp_service.service import WhatsAppService
        from apps.extensions import db
        
        data = request.get_json(silent=True) or {}
        campaign_name = data.get('campaign_name', '').strip()
        target_category = data.get('target_category', 'all')
        message_text = data.get('message_text', '').strip()
        
        if not campaign_name or not message_text:
            return jsonify({"success": False, "error": "اسم الحملة والرسالة مطلوبان"}), 400
        
        # جلب جهات الاتصال حسب الفئة
        query = WhatsAppCustomerContact.query
        if target_category != 'all':
            query = query.filter_by(category=target_category)
        
        contacts = query.all()
        
        if not contacts:
            return jsonify({"success": False, "error": "لا توجد جهات اتصال في هذه الفئة"}), 400
        
        wa_service = WhatsAppService()
        
        sent = 0
        failed = 0
        
        for contact in contacts:
            try:
                msg = message_text.replace('{name}', contact.name or '')
                msg = msg.replace('{phone}', contact.phone or '')
                msg = msg.replace('{company}', getattr(contact, 'company', '') or '')
                msg = msg.replace('{city}', getattr(contact, 'city', '') or '')
                
                wa_service.send_message(contact.phone, msg)
                sent += 1
            except Exception as e:
                logger.error(f"فشل إرسال رسالة إلى {contact.phone}: {str(e)}")
                failed += 1
        
        return jsonify({
            "success": True,
            "sent": sent,
            "failed": failed,
            "total": len(contacts),
            "campaign": campaign_name,
            "target_category": target_category,
            "message": f"تم إرسال الحملة '{campaign_name}' إلى {sent} جهة اتصال"
        }), 200
    except Exception as e:
        logger.error(f"خطأ في إرسال الحملة: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


# =========================================================================
# 6. توليد رسالة بالذكاء الاصطناعي
# =========================================================================

@contacts_bulk_bp.route('/api/generate-message', methods=['POST'])
def generate_message_api():
    """توليد رسالة تسويقية بالذكاء الاصطناعي (Gemini)"""
    try:
        from apps.whatsapp_service.service import WhatsAppService
        wa_service = WhatsAppService()
        
        data = request.get_json(silent=True) or {}
        category = data.get('category', 'customers')
        
        # أسماء الفئات بالعربية
        category_names = {
            'customers': 'العملاء',
            'merchants': 'التجار',
            'suppliers': 'الموردين',
            'marketers': 'المسوقين',
            'all': 'جميع الفئات'
        }
        
        category_name = category_names.get(category, 'العملاء')
        
        # استخدام Gemini API لتوليد النص
        prompt = f"""
        أنت مساعد تسويقي متخصص في سوق محجوب أونلاين.
        قم بكتابة رسالة تسويقية احترافية وجذابة لـ {category_name}،
        تحتوي على:
        1. تحية
        2. عرض مميز
        3. دعوة للشراء أو التواصل
        4. شكر وتقدير
        استخدم المتغيرات: {{name}}, {{company}}, {{city}}
        كن مختصراً وجذاباً، لا تتجاوز 150 كلمة.
        """
        
        response = wa_service._generate_gemini_reply(prompt)
        
        return jsonify({"success": True, "text": response}), 200
    except Exception as e:
        logger.error(f"خطأ في توليد الرسالة: {str(e)}")
        
        # رسائل افتراضية في حالة فشل Gemini
        default_messages = {
            'customers': """مرحباً {name}،

يسعدنا في سوق محجوب أونلاين أن نقدم لك عروضاً حصرية على أحدث المنتجات! 🛍️

🔹 خصم 20% على أول طلب
🔹 شحن مجاني للطلبات فوق 200 ريال
🔹 توصيل سريع لجميع المناطق

استخدم كود الخصم: WELCOME20

🌟 ننتظرك في سوق محجوب أونلاين!""",
            
            'merchants': """عزيزي التاجر {name}،

نقدم لك في سوق محجوب أونلاين فرصة لتوسيع أعمالك! 📈

🔹 منصة لعرض منتجاتك
🔹 وصول لأكثر من 10,000 عميل
🔹 أدوات تسويق متقدمة

انضم الآن واستفد من خصم 30% على الاشتراك السنوي!

نحن هنا لدعم نجاحك! 🚀""",
            
            'suppliers': """مرحباً {name}،

سوق محجوب أونلاين يبحث عن موردين جدد! 🤝

🔹 منصة موثوقة للبيع بالجملة
🔹 عقود طويلة الأجل
🔹 دعم لوجستي متكامل

سجل منتجاتك الآن واستفد من قاعدة عملائنا الواسعة!

نتطلع لشراكتك! 💪""",
            
            'marketers': """مرحباً {name}،

سوق محجوب أونلاين يفتح باب التسويق بالعمولة! 💰

🔹 عمولة تصل إلى 20% على كل عملية بيع
🔹 مواد تسويقية جاهزة
🔹 تقارير وتحليلات متقدمة

انضم إلى فريق المسوقين وابدأ في تحقيق أرباحك اليوم!

معاً نحو النجاح! 🌟"""
        }
        
        default_text = default_messages.get(category, default_messages['customers'])
        return jsonify({"success": True, "text": default_text}), 200


# =========================================================================
# 7. تصفير عداد الرسائل غير المقروءة
# =========================================================================

@contacts_bulk_bp.route('/api/contacts/<phone>/read', methods=['POST'])
def mark_contact_read_by_phone(phone):
    """تصفير عداد الرسائل غير المقروءة لمستخدم معين"""
    try:
        from apps.whatsapp_service.service import WhatsAppService
        wa_service = WhatsAppService()
        
        result = wa_service.mark_contact_as_read(phone)
        return jsonify(result), 200
    except Exception as e:
        logger.error(f"خطأ في تصفير العداد: {str(e)}")
        return jsonify({"error": str(e)}), 500


# =========================================================================
# 8. إحصائيات جهات الاتصال
# =========================================================================

@contacts_bulk_bp.route('/api/contacts/stats', methods=['GET'])
def get_contacts_stats_api():
    """جلب إحصائيات جهات الاتصال"""
    try:
        from apps.models.whatsapp_models import WhatsAppCustomerContact
        from apps.extensions import db
        
        total = WhatsAppCustomerContact.query.count()
        customers = WhatsAppCustomerContact.query.filter_by(category='customers').count()
        merchants = WhatsAppCustomerContact.query.filter_by(category='merchants').count()
        suppliers = WhatsAppCustomerContact.query.filter_by(category='suppliers').count()
        marketers = WhatsAppCustomerContact.query.filter_by(category='marketers').count()
        
        return jsonify({
            "success": True,
            "stats": {
                "total": total,
                "customers": customers,
                "merchants": merchants,
                "suppliers": suppliers,
                "marketers": marketers
            }
        }), 200
    except Exception as e:
        logger.error(f"خطأ في جلب الإحصائيات: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500
